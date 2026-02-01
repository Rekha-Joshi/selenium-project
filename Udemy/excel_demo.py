#from unittest.mock import _NameArgsKwargs
from selenium import webdriver
from openpyxl import load_workbook, workbook

#load existing workbook
wb = load_workbook("test_data.xlsx")

#select active sheet or by name
sheet = wb.active
# or sheet = wb["Products"]

print(f"Current active sheet name: {sheet.title}")
print(f"Total rows: {sheet.max_row}")
print(f"Total columns: {sheet.max_column}")
print(f"All sheets: {wb.sheetnames}")

#iterate rows
print("\nProducts: Using iter_rows")
for row in sheet.iter_rows(values_only=True):
    print(row)

#loop through all rows and columns
print("\nProducts: Using for loop with max_row and max_column")
for row in range(1, sheet.max_row+1):
    for column in range(1, sheet.max_column):
        cell_value = sheet.cell(row=row, column=column).value
        print(cell_value, end=" | ")
        #Python prints the value and then automatically adds a new line (\n) after it.
        #The end parameter of print() changes what happens at the end of each print. it will replace \n with |
    print() #print new line after each row

#creating and writing into into sheet. when you run this file again, it will create another expenses1, 2, and so on
new_sheet = wb.create_sheet("Expenses")

#write data to it
new_sheet["A1"]="Item"
new_sheet["B1"]="Amount"

#append rows
new_sheet.append(["Rekha", 250])
new_sheet.append(["Saisha", 115])

#save worbook
wb.save("test_data.xlsx")
print(f"New sheet list: {wb.sheetnames}")

#print Expenses
sheet = wb["Expenses"]
print("\nExpenses:")
for row in sheet.iter_rows(values_only=True):
    print(row)

#delete specific sheet and keep other
#sheet_to_delete = wb["Expenses2"]
#wb.remove(sheet_to_delete)
print("=====")
keep_sheets = ["Products", "Vendors", "Notes"]
for sheet_name in list(wb.sheetnames): #ist(wb.sheetnames) → makes a new independent list copy.
    if sheet_name not in keep_sheets:
        wb.remove(wb[sheet_name])
        print(f"🗑️ Deleted sheet: {sheet_name}")
    else:
        print(f"✅ Kept sheet: {sheet_name}")

# Save changes
wb.save("test_data.xlsx")


