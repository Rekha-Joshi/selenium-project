from logging import exception
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

from openpyxl import load_workbook, workbook

#global variables
downloaded_file = "/Users/rekhajoshi/Downloads/download.xlsx"
fruit_to_update="Papaya"
new_price = 640

def update_price(file_path, fruit_to_update, new_price):
    # here the logic will come to update the excel
    wb = load_workbook(file_path)
    sheet = wb.active

    # Initialize first (avoid UnboundLocalError)
    fruit_col = None
    price_col = None

    #Before updating, we should know which column has fruit and which one has price
    for col in sheet.iter_cols(min_row=1, max_row=1):
        header = str(col[0].value).strip().lower() if col[0].value else "" #col[0].value reads the name of the header
        if header == 'fruit_name':
            fruit_col = col[0].column #.column gives the column index
        if header == 'price':
            price_col = col[0].column
    if not fruit_col or not price_col:
        raise Exception("Couldn't find fruit or price column")
    #we now know the column number of fruit and price, we will go through all rows.
    for row in range(2, sheet.max_row+1):
        if sheet.cell(row, fruit_col).value.strip().lower() == fruit_to_update.lower():
            sheet.cell(row, price_col).value = new_price
            break
    else:
        print(f"{fruit_to_update} not found")
    #save workbook
    wb.save(downloaded_file)
    wb.close()

#define driver
driver = webdriver.Chrome()
driver.maximize_window()

#define explicit wait
wait = WebDriverWait(driver,5)

#get url
driver.get("https://rahulshettyacademy.com/upload-download-test/")

#click on download
download_btn = wait.until(EC.element_to_be_clickable((By.ID, "downloadButton")))
download_btn.click()

# --- Step 1: find Price column ID ---
price_header = driver.find_element(By.XPATH, "//div[normalize-space(.)='Price']")
col_id = price_header.get_attribute("data-column-id")

# --- Step 2: get OLD price from web ---
fruit_price_xpath = f"//div[text()='{fruit_to_update}']/parent::div/parent::div/div[@id='cell-{col_id}-undefined']"
old_price_web = driver.find_element(By.XPATH, fruit_price_xpath).text
print(f"Before update '{fruit_to_update}' price is {old_price_web}")

# --- Step 3: update Excel file dynamically ---
update_price(downloaded_file, fruit_to_update, new_price)

# --- Step 4: re-upload Excel ---
driver.find_element(By.ID, "fileinput").send_keys(downloaded_file)

#wait for the success toast and assert the success
toast = wait.until(EC.visibility_of_element_located((By.XPATH, "//div[@class='Toastify__toast-body']/div[2]")))
print(toast.text)
assert "Successfully" in toast.text

## --- Step 5: get NEW price from web ---
new_price_web = driver.find_element(By.XPATH, fruit_price_xpath).text
print(f"New price on web: {new_price_web}")

# --- Step 6: verify ---
assert int(new_price_web) == new_price
print(f"Price updated successfully: {old_price_web} → {new_price_web}")
driver.quit()
